"""
Production script to extract metadata from structural steel drawing PDFs and generate Excel transmittal.

This script processes PDFs with filenames in the format:
  "DRAWING_NO - TITLE - Rev X.pdf"
  
Example: "2AL46 - LINTEL ANGLE - Rev 0.pdf"

Usage: 
  python generate_transmittal.py [folder_path]
  
If no folder is specified, you'll be prompted to enter one.
"""
import os
import sys
import re
from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.extraction_parser import PDFParser

def extract_from_filename(filename):
    """
    Extracts drawing metadata from filename.
    Expected format: "DRAWING_NO - TITLE - Rev X.pdf"
    
    Returns:
        dict with drawing_no, description, revision_no
    """
    # Remove .pdf extension
    name = os.path.splitext(filename)[0]
    
    # Try to match pattern: "DRAWING_NO - TITLE - Rev X"
    match = re.match(r'^([A-Z0-9]+)\s*-\s*(.+?)\s*-\s*Rev\s*([A-Z0-9]+)', name, re.IGNORECASE)
    
    if match:
        return {
            'drawing_no': match.group(1).strip(),
            'description': match.group(2).strip(),
            'revision_no': match.group(3).strip()
        }
    
    # Fallback: try simpler pattern "DRAWING_NO - TITLE"
    match2 = re.match(r'^([A-Z0-9]+)\s*-\s*(.+)', name)
    if match2:
        return {
            'drawing_no': match2.group(1).strip(),
            'description': match2.group(2).strip(),
            'revision_no': '0'
        }
    
    # Last resort: use whole filename as drawing number
    return {
        'drawing_no': name,
        'description': '',
        'revision_no': '0'
    }

def create_transmittal_excel(drawings_data, output_path, project_name="Project"):
    """
    Creates an Excel transmittal file matching the required format.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drawing Transmittal"
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    
    # Add Logo
    try:
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "caldim_logo.png")
        if os.path.exists(logo_path):
            img = Image(logo_path)
            # Resize image to fit in rows 1-5 (approx 90-100 pixels height)
            # Original is 1024x400. Let's make it height=80.
            # Allow aspect ratio preservation
            h_ratio = 80 / img.height
            new_width = int(img.width * h_ratio)
            
            img.height = 80
            img.width = new_width
            
            # Place the logo at C1 (approximate center)
            ws.add_image(img, 'C1')
        else:
            print(f"Warning: Logo not found at {logo_path}")
    except Exception as e:
        print(f"Warning: Could not add logo: {e}")

    # Header section
    ws['A6'] = "Customer Name:"
    ws['A7'] = f"Project Name: {project_name}"
    ws['A8'] = "Customer Project No:"
    ws['A9'] = f"TRANSMITTAL #{datetime.now().strftime('%Y%m%d')}"
    
    # Column headers (row 10)
    headers = ["Sl. No.", "Sheet No.", "Drawing Title", "Revision Mark", "Date", "Remarks"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Section header (row 11) - "SHOP DRAWINGS"
    current_row = 11
    ws.merge_cells(f'A{current_row}:F{current_row}')
    section_cell = ws.cell(row=current_row, column=1)
    section_cell.value = "SHOP DRAWINGS"
    section_cell.font = Font(bold=True)
    section_cell.alignment = Alignment(horizontal='center', vertical='center')
    section_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    current_row += 1
    
    # Data rows
    for idx, drawing in enumerate(drawings_data, 1):
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=drawing.get('drawing_no', ''))
        ws.cell(row=current_row, column=3, value=drawing.get('description', ''))
        ws.cell(row=current_row, column=4, value=drawing.get('revision_no', '0'))
        ws.cell(row=current_row, column=5, value=drawing.get('date', ''))
        ws.cell(row=current_row, column=6, value=drawing.get('remarks', 'For Fabrication'))
        
        # Add borders
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        current_row += 1
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Transmittal Excel saved to: {output_path}")
    return output_path

def process_pdfs(folder_path):
    """
    Process all PDFs in the folder and extract metadata.
    """
    parser = PDFParser()
    drawings = []
    
    # Find all PDFs recursively
    pdf_files = sorted(list(Path(folder_path).rglob("*.pdf")))
    
    if not pdf_files:
        print(f"No PDF files found in: {folder_path}")
        return []
    
    print(f"Found {len(pdf_files)} PDF file(s)\n")
    
    for pdf_file in pdf_files:
        print(f"  Processing: {pdf_file.name}...", end=" ")
        
        try:
            # Extract from filename (primary method)
            file_data = extract_from_filename(pdf_file.name)
            
            # Try to extract date and remarks from PDF
            try:
                pdf_metadata = parser.extract_metadata(str(pdf_file))
                if pdf_metadata and not pdf_metadata.get("is_transmittal"):
                    # Use PDF date if available
                    if pdf_metadata.get('date'):
                        file_data['date'] = pdf_metadata['date']
                    # Use PDF remarks if available
                    if pdf_metadata.get('remarks'):
                        file_data['remarks'] = pdf_metadata['remarks']
            except:
                pass  # If PDF extraction fails, use filename data only
            
            # Set defaults if not present
            if 'date' not in file_data:
                file_data['date'] = datetime.now().strftime("%m/%d/%Y")
            if 'remarks' not in file_data:
                file_data['remarks'] = "For Fabrication"
            
            drawings.append(file_data)
            print(f"✓ {file_data['drawing_no']}")
            
        except Exception as e:
            print(f"✗ Error: {e}")
    
    return drawings

if __name__ == "__main__":
    print("="*80)
    print("STEEL DRAWING TRANSMITTAL GENERATOR")
    print("="*80)
    print()
    
    # Get folder path
    if len(sys.argv) > 1:
        folder_path = sys.argv[1]
    else:
        folder_path = input("Enter the folder path containing PDF files: ").strip().strip('"')
    
    if not os.path.exists(folder_path):
        print(f"\n✗ ERROR: Folder not found: {folder_path}")
        sys.exit(1)
    
    if not os.path.isdir(folder_path):
        print(f"\n✗ ERROR: Not a directory: {folder_path}")
        sys.exit(1)
    
    print(f"Folder: {folder_path}\n")
    
    # Process PDFs
    drawings = process_pdfs(folder_path)
    
    if not drawings:
        print("\n✗ No drawings extracted!")
        sys.exit(1)
    
    # Generate Excel
    print(f"\n{'='*80}")
    print(f"SUMMARY: Extracted {len(drawings)} drawing(s)")
    print(f"{'='*80}\n")
    
    output_path = os.path.join(os.path.dirname(folder_path), "Transmittal_Output.xlsx")
    create_transmittal_excel(drawings, output_path)
    
    print(f"\nFirst 10 drawings:")
    for i, d in enumerate(drawings[:10], 1):
        print(f"  {i}. {d['drawing_no']} - {d['description']} (Rev {d['revision_no']})")
    
    if len(drawings) > 10:
        print(f"  ... and {len(drawings) - 10} more")
    
    print(f"\n{'='*80}")
    print(f"✓ COMPLETE! Excel file: {output_path}")
    print(f"{'='*80}")
