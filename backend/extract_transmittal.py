"""
Script to extract metadata from structural steel drawing PDFs and generate Excel transmittal.
Usage: python extract_transmittal.py [folder_path]
"""
import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.extraction_parser import PDFParser

def create_transmittal_excel(drawings_data, output_path, project_name="Project"):
    """
    Creates an Excel transmittal file matching the sample format.
    
    Args:
        drawings_data: List of drawing metadata dicts
        output_path: Path to save the Excel file
        project_name: Project name for the header
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drawing Transmittal"
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Header section (rows 1-9)
    ws['A6'] = "Customer Name:"
    ws['A7'] = f"Project Name: {project_name}"
    ws['A8'] = f"Customer Project No: "
    ws['A9'] = f"TRANSMITTAL #{datetime.now().strftime('%Y%m%d')}"
    
    # Column headers (row 10)
    headers = ["Sl. No.", "Sheet No.", "Drawing Title", "Revision Mark", "Date", "Remarks"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Section header (row 11) - "SHOP DRAWINGS"
    current_row = 11
    ws.merge_cells(f'A{current_row}:F{current_row}')
    section_cell = ws.cell(row=current_row, column=1)
    section_cell.value = "SHOP DRAWINGS"
    section_cell.font = Font(bold=True)
    section_cell.alignment = Alignment(horizontal='center', vertical='center')
    section_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    current_row += 1
    
    # Data rows
    for idx, drawing in enumerate(drawings_data, 1):
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=drawing.get('drawing_no', ''))
        ws.cell(row=current_row, column=3, value=drawing.get('description', ''))
        ws.cell(row=current_row, column=4, value=drawing.get('revision_no', '0'))
        ws.cell(row=current_row, column=5, value=drawing.get('date', ''))
        ws.cell(row=current_row, column=6, value=drawing.get('remarks', 'For Fabrication'))
        
        # Add borders to all cells
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        current_row += 1
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Transmittal Excel saved to: {output_path}")

def extract_from_folder(folder_path, recursive=False):
    """
    Extracts metadata from all PDFs in a folder.
    
    Args:
        folder_path: Path to folder containing PDF files
        recursive: If True, search subdirectories
        
    Returns:
        List of drawing metadata dicts
    """
    parser = PDFParser()
    drawings = []
    
    if recursive:
        pdf_files = list(Path(folder_path).rglob("*.pdf"))
    else:
        pdf_files = list(Path(folder_path).glob("*.pdf"))
    
    print(f"Found {len(pdf_files)} PDF files in {folder_path}")
    
    for pdf_file in pdf_files:
        print(f"  Processing: {pdf_file.name}...", end=" ")
        try:
            metadata = parser.extract_metadata(str(pdf_file))
            
            if metadata:
                # Check if it's a transmittal (contains multiple drawings)
                if metadata.get("is_transmittal"):
                    drawings.extend(metadata.get("drawings", []))
                    print(f"✓ Transmittal with {len(metadata.get('drawings', []))} drawings")
                else:
                    drawings.append(metadata)
                    print(f"✓ {metadata.get('drawing_no')} - {metadata.get('description')}")
            else:
                print("✗ No metadata")
                    
        except Exception as e:
            print(f"✗ Error: {e}")
    
    return drawings

if __name__ == "__main__":
    # Get folder path from command line or use default
    if len(sys.argv) > 1:
        folder_path = sys.argv[1]
    else:
        # Default: use the rr subfolder which likely has the drawings
        folder_path = r"d:\steel(3)\steel\storage\rr"
    
    output_path = r"d:\steel(3)\steel\backend\Transmittal_Output.xlsx"
    
    print("="*80)
    print("STEEL DRAWING METADATA EXTRACTION TO EXCEL TRANSMITTAL")
    print("="*80)
    print(f"Folder: {folder_path}")
    print()
    
    if not os.path.exists(folder_path):
        print(f"ERROR: Folder not found: {folder_path}")
        print("\nUsage: python extract_transmittal.py [folder_path]")
        sys.exit(1)
    
    drawings = extract_from_folder(folder_path)
    
    print(f"\n{'='*80}")
    print(f"SUMMARY: Extracted {len(drawings)} drawing(s)")
    print(f"{'='*80}")
    
    if drawings:
        # Get project name from first drawing
        project_name = drawings[0].get('project_name', 'Project')
        
        create_transmittal_excel(drawings, output_path, project_name)
        
        print(f"\nFirst 10 drawings:")
        for i, d in enumerate(drawings[:10], 1):
            print(f"  {i}. {d.get('drawing_no')} - {d.get('description')} (Rev {d.get('revision_no')})")
        
        if len(drawings) > 10:
            print(f"  ... and {len(drawings) - 10} more")
    else:
        print("\n✗ No drawings extracted!")
        print("Please check that the folder contains PDF files.")
