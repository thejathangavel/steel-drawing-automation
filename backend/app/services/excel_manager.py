import openpyxl
import re
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class ExcelManager:
    def __init__(self):
        pass

    @staticmethod
    def _natural_keys(text):
        """
        alist.sort(key=natural_keys) sorts in human order
        """
        if not isinstance(text, str):
            text = str(text)
        return [ int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text) ]

    def create_transmittal(self, drawings, output_path, project_name="Project"):
        """
        Creates a transmittal excel file matching the sample:
        Rows 6-9: Metadata
        Row 10: Headers
        Row 11+: Data
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transmittal"
        
        # Styles
        title_font = Font(bold=True, size=11, name='Calibri')
        header_font = Font(bold=True, size=10, name='Calibri')
        normal_font = Font(size=10, name='Calibri')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        today_date = datetime.now().strftime("%m/%d/%Y")

        # --- Metadata Section ---
        # Add Logo
        try:
            # Navigate up two levels from app/services: services -> app -> backend
            # Actually __file__ is inside services/excel_manager.py
            # dirname(__file__) is .../services
            # dirname(...) is .../app
            # dirname(...) is .../backend
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            logo_path = os.path.join(base_dir, "assets", "caldim_logo.png")
            
            if os.path.exists(logo_path):
                img = Image(logo_path)
                
                # Resize image (height=80 pixels, keep aspect ratio)
                h_target = 80
                h_ratio = h_target / img.height
                img.height = h_target
                img.width = int(img.width * h_ratio)
                
                ws.add_image(img, 'C1')
            else:
                 print(f"Warning: Logo not found at {logo_path}")
        except Exception as e:
             print(f"Warning: Could not add logo: {e}")

        # R6: Customer Name
        ws['A6'] = "Customer Name: "
        ws['A6'].font = normal_font
        ws.merge_cells('A6:F6')

        # R7: Project Name
        ws['A7'] = f"Project Name: {project_name}"
        ws['A7'].font = normal_font
        ws.merge_cells('A7:F7')

        # R8: Customer Project No
        ws['A8'] = "Customer Project No: 24050" 
        ws['A8'].font = normal_font
        ws.merge_cells('A8:F8')
        
        # R9: Transmittal Title
        ws['A9'] = f"TRANSMITTAL #62 DT.{today_date}"
        ws['A9'].font = title_font
        ws['A9'].alignment = center_align
        ws.merge_cells('A9:F9')
        # Border for title
        for col in range(1, 7):
            ws.cell(row=9, column=col).border = thin_border

        # --- Header Section (Row 10) ---
        headers = ["Sl. No.", "Sheet No.", "Drawing Title", "Revision Mark", "Date", "Remarks"]
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_idx, value=text)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            # Widths
            if col_idx == 1: ws.column_dimensions['A'].width = 8
            elif col_idx == 2: ws.column_dimensions['B'].width = 15
            elif col_idx == 3: ws.column_dimensions['C'].width = 50
            elif col_idx == 4: ws.column_dimensions['D'].width = 15
            elif col_idx == 5: ws.column_dimensions['E'].width = 15
            elif col_idx == 6: ws.column_dimensions['F'].width = 30

        current_row = 11

        # Categorize
        e_sheets = []
        shop_drawings = []
        part_drawings = []
        
        # Helper to categorize based on drawing data
        for d in drawings:
            # d is expected to be a dict
            d_type = str(d.get('drawing_type', '')).upper()
            d_no = str(d.get('drawing_no', '')).upper()
            description = str(d.get('description', '')).upper()
            
            # 1. Erection Drawings
            if d_type == 'ERECTION' or d_no.startswith('E') or 'ERECTION' in description:
                e_sheets.append(d)
            # 2. Shop Drawings (Explicit Type OR AL heuristic)
            elif d_type == 'SHOP' or 'AL' in d_no:
                 shop_drawings.append(d)
            # 3. Part Drawings (Everything else - 1B, 1C, 1PL, etc.)
            else:
                part_drawings.append(d)

        # Sort
        e_sheets.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))
        shop_drawings.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))
        part_drawings.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))

        def write_section(title, items, start_sl):
            nonlocal current_row
            # Section Header
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            # Yellow fill for E SHEET / SHOP DRAWINGS as per sample
            cell.fill = yellow_fill 
            for col in range(2, 7): ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1
            
            for i, item in enumerate(items, start_sl):
                # Retrieve raw values
                raw_dwg_no = str(item.get('drawing_no', '')).strip()
                raw_rev = str(item.get('revision_no', '')).strip()
                raw_remarks = str(item.get('remarks', '')).strip()

                # --- FIX 2: Sheet No Mapping (Remove Dates) ---
                # Check if drawing_no is a date (e.g. 06/26/2025)
                # If so, it's invalid for Sheet No.
                if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', raw_dwg_no):
                    # Invalid drawing number (it's a date). Clear it.
                    raw_dwg_no = ""

                # --- FIX 1: Revision Mark Formatting ---
                # Clean generic text from Revision Mark
                clean_rev = raw_rev
                for text_to_remove in ["For Approval", "For Fabrication", "Rev", "rev", "REV", "Revision", "mark", "."]:
                    clean_rev = clean_rev.replace(text_to_remove, "").strip()
                
                # Logic based on Remarks
                upper_remarks = raw_remarks.upper()
                
                if "FOR APPROVAL" in upper_remarks:
                    # Must be Alphabetic (A, B, C...)
                    # If it's numeric 0, convert to A
                    if clean_rev == '0':
                        clean_rev = 'A'
                    # If strictly numeric and not 0, map logic could be complex, but let's assume 0->A is main case
                    # If empty, default to A
                    if not clean_rev:
                        clean_rev = 'A'
                    # Ensure no digits left if possible, or just accept whatever clean state
                    
                elif "FOR FABRICATION" in upper_remarks:
                    # Must be Numeric (0, 1, 2...)
                    # If it's 'A', convert to 0? Or just stripping non-digits?
                    if clean_rev.upper() == 'A':
                        clean_rev = '0'
                    
                    # Try to ensure it's numeric
                    if not clean_rev.isdigit():
                        # Extract digits only
                        digits = "".join(c for c in clean_rev if c.isdigit())
                        if digits:
                            clean_rev = digits
                        else:
                            clean_rev = '0' # Default for fabrication

                # Sl No
                c = ws.cell(row=current_row, column=1, value=i)
                c.font = normal_font
                c.alignment = center_align
                c.border = thin_border

                # Sheet No (Drawing No)
                c = ws.cell(row=current_row, column=2, value=raw_dwg_no)
                c.font = normal_font
                c.alignment = center_align
                c.border = thin_border

                # Title
                c = ws.cell(row=current_row, column=3, value=item.get('description', ''))
                c.font = normal_font
                c.alignment = left_align
                c.border = thin_border

                # Rev
                c = ws.cell(row=current_row, column=4, value=clean_rev)
                c.font = normal_font
                c.alignment = center_align
                c.border = thin_border

                # Date
                val_date = item.get('date', '')
                try:
                     # Try to format date nicely
                    if '-' in val_date: # 2023-10-10
                         val_date = datetime.strptime(val_date, "%Y-%m-%d").strftime("%d/%m/%Y")
                except:
                    pass
                c = ws.cell(row=current_row, column=5, value=val_date)
                c.font = normal_font
                c.alignment = center_align
                c.border = thin_border

                # Remarks - Standardize and check for HOLD
                final_remarks = raw_remarks
                
                if "FOR APPROVAL" in upper_remarks:
                    final_remarks = "For Approval"
                elif "FOR FABRICATION" in upper_remarks:
                    final_remarks = "For Fabrication"
                    
                if "HOLD" in upper_remarks:
                    final_remarks = f"{final_remarks} (HOLD)"
                
                c = ws.cell(row=current_row, column=6, value=final_remarks)
                c.font = normal_font
                c.alignment = left_align
                c.border = thin_border
                
                current_row += 1
            return start_sl + len(items)

        next_sl = 1
        if e_sheets:
            next_sl = write_section("E SHEET", e_sheets, 1)

        # Reset SL for next section? Usually yes for new Category
        if shop_drawings:
            write_section("SHOP DRAWINGS", shop_drawings, 1)

        if part_drawings:
             write_section("PART DRAWINGS", part_drawings, 1)

        # Footer
        row = current_row
        ws[f'A{row}'] = "Prepared By : SMS"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        
        ws[f'C{row}'] = "PSR"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'C{row}'].alignment = center_align
        
        ws[f'D{row}'] = "Sent By : HAR"
        ws[f'D{row}'].font = Font(bold=True)
        ws.merge_cells(f'D{row}:F{row}')

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

        wb.save(output_path)
        return output_path

    def update_drawing_log(self, drawing_log_data, output_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Drawing Log"
        
        # Styles
        header_font = Font(bold=True, size=10, name='Calibri')
        normal_font = Font(size=10, name='Calibri')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Metadata
        
        # Add Logo
        try:
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            logo_path = os.path.join(base_dir, "assets", "caldim_logo.png")
            
            if os.path.exists(logo_path):
                img = Image(logo_path)
                h_target = 80
                h_ratio = h_target / img.height
                img.height = h_target
                img.width = int(img.width * h_ratio)
                ws.add_image(img, 'C1')
            else:
                 print(f"Warning: Logo not found at {logo_path}")
        except Exception as e:
             print(f"Warning: Could not add logo: {e}")

        # Shifted Metadata (Start Row 6)
        ws['A6'] = "Project Name: GSC Juniper RBP Phase II"
        ws.merge_cells('A6:K6') # Wide merge
        ws['A6'].font = header_font
        
        ws['A7'] = "OUTGOING DRAWING LOG SHEET"
        ws.merge_cells('A7:K7')
        ws['A7'].alignment = center_align
        ws['A7'].font = Font(bold=True, size=12)

        # Determine Max Revisions for Headers
        max_alpha = 0
        max_num = 0
        
        for item in drawing_log_data:
            # Backfill revisions if empty but we have current data
            revs = item.get('revisions', [])
            
            # Use current revision/date/status if not in history
            curr_rev = str(item.get('revision_no', '')).strip()
            curr_date = item.get('date', '')
            status = str(item.get('remarks', '') or item.get('status', '')).upper()

            # Infer revision from status if missing
            # Infer or Correct revision from status
            # Logic: 
            # 1. If 'APPROVAL' is in remarks, Rev MUST be Alpha. If it's numeric/empty, set to 'A'.
            # 2. If 'FABRICATION' is in remarks, Rev MUST be Numeric. If it's alpha/empty, set to '0'.
            
            is_alpha_rev = curr_rev.replace('+','').isalpha() if curr_rev else False
            is_numeric_rev = curr_rev.isdigit() if curr_rev else False
            
            if 'APPROVAL' in status:
                # Force Alpha.
                # 1. If current rev is not alpha, change it to A.
                if not curr_rev or not is_alpha_rev:
                    curr_rev = "A"
                # CLEANUP REMOVED: Do not remove numeric revisions. We want to keep history (Rev 0) even if current is Approval.
                
            elif 'FABRICATION' in status:
                 # Force Numeric.
                 # 1. If current rev is alpha, change it to 0.
                 if not curr_rev or is_alpha_rev:
                    curr_rev = "0"
                 # CLEANUP REMOVED: Do not remove alpha revisions.

            # Update item with inferred revision for consistency
            if curr_rev:
                item['revision_no'] = curr_rev
            
            # Rebind clean history
            item['revisions'] = revs
            
            # Check if this current revision is already in history
            rev_exists = any(str(r['rev']).strip().upper() == curr_rev.upper() for r in revs)
            
            if not rev_exists and curr_rev:
                 # Add it to local list for calculation purposes (and later display)
                 revs.append({'rev': curr_rev, 'date': curr_date})
                 item['revisions'] = revs
            
            if '1AL4' in str(item.get('drawing_no', '')):
                print(f"DEBUG 1AL4: Date='{curr_date}', Status='{status}', Revs={revs}, MaxAlpha={max_alpha}")
            
            alphas = [r for r in revs if str(r['rev']).upper() in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ']
            nums = [r for r in revs if str(r['rev']).isdigit()]
            if len(alphas) > max_alpha: max_alpha = len(alphas)
            if len(nums) > max_num: max_num = len(nums)
            
        # --- NEW LOGIC: Common Date Inference for Approval ---
        # 1. Collect all dates to find Mode
        all_dates = []
        for item in drawing_log_data:
            d = item.get('date', '')
            # Relaxed check: Accept any non-empty string that looks like a date (contains numbers and separators)
            # YYYY-MM-DD, DD-MM-YYYY, MM/DD/YYYY, etc.
            d_str = str(d).strip()
            if d_str and len(d_str) >= 8 and (re.search(r'\d', d_str)) and (re.search(r'[-\./]', d_str)):
                all_dates.append(d_str)
        
        common_date = None
        if all_dates:
            # Find mode (most frequent date)
            from collections import Counter
            common_date = Counter(all_dates).most_common(1)[0][0]
            print(f"DEBUG: Common Batch Date inferred as: {common_date}")
        else:
            # Fallback to Today if absolutely no dates found in batch
            common_date = datetime.now().strftime("%m-%d-%Y")
            print(f"DEBUG: No dates in batch. Fallback to Today: {common_date}")
            
        # 2. Backfill Approval items missing date
        if common_date:
            for item in drawing_log_data:
                # Check if it is an Approval item (Rev A inferred)
                revs = item.get('revisions', [])
                status = str(item.get('remarks', '')).upper()
                
                if "APPROVAL" in status:
                     # Check if we have a Rev A entry with a date
                    has_date_for_a = False
                    for r in revs:
                        if str(r['rev']).upper() == 'A' and r.get('date'):
                            has_date_for_a = True
                            break
                    
                    if not has_date_for_a:
                        # Force inject Rev A with Common Date
                        # Remove any existing bad Rev A
                        revs = [r for r in revs if str(r['rev']).upper() != 'A']
                        revs.append({'rev': 'A', 'date': common_date})
                        item['revisions'] = revs
                        # Also update main item date for display if needed
                        if not item.get('date'): item['date'] = common_date

                # Note: We do NOT do this for Fabrication (Rev 0) as fabrication strictly requires a date on drawing usually.
                # But user only asked for Approval inference.

            
        # Min widths from sample - REMOVED FORCED MINIMUMS
        # max_alpha = max(max_alpha, 2)
        # max_num = max(max_num, 4)
        
        # --- Row 8 Headers ---
        # A8: Sl No, B8: Sheet No, C8: Drawing Title
        ws.cell(row=8, column=1, value="Sl. No.").border = thin_border
        ws.merge_cells('A8:A9')
        ws.cell(row=8, column=1).alignment = center_align

        ws.cell(row=8, column=2, value="Sheet No.").border = thin_border
        ws.merge_cells('B8:B9')
        ws.cell(row=8, column=2).alignment = center_align

        ws.cell(row=8, column=3, value="Drawing Title").border = thin_border
        ws.merge_cells('C8:C9')
        ws.cell(row=8, column=3).alignment = center_align
        
        # Dynamic Column Tracking
        current_col = 4
        
        # Sent For Approval (Columns D to ...)
        start_approval = current_col
        if max_alpha > 0:
            end_approval = start_approval + max_alpha - 1
            ws.cell(row=8, column=start_approval, value="Sent for Approval")
            ws.merge_cells(start_row=8, start_column=start_approval, end_row=8, end_column=end_approval)
            ws.cell(row=8, column=start_approval).alignment = center_align
            for c in range(start_approval, end_approval+1): ws.cell(row=8, column=c).border = thin_border
            current_col += max_alpha
        else:
            # If no alpha revisions, no columns added
            pass
        
        # Sent For Fabrication
        start_fab = current_col
        if max_num > 0:
            end_fab = start_fab + max_num - 1
            ws.cell(row=8, column=start_fab, value="Sent for Fabrication")
            ws.merge_cells(start_row=8, start_column=start_fab, end_row=8, end_column=end_fab)
            ws.cell(row=8, column=start_fab).alignment = center_align
            for c in range(start_fab, end_fab+1): ws.cell(row=8, column=c).border = thin_border
            current_col += max_num
        else:
            pass
        
        # Remarks & Sequence
        col_remarks = current_col
        col_seq = current_col + 1
        
        ws.cell(row=8, column=col_remarks, value="Remarks").border = thin_border
        ws.merge_cells(start_row=8, start_column=col_remarks, end_row=9, end_column=col_remarks)
        ws.cell(row=8, column=col_remarks).alignment = center_align
        
        ws.cell(row=8, column=col_seq, value="Sequence/ Area").border = thin_border
        ws.merge_cells(start_row=8, start_column=col_seq, end_row=9, end_column=col_seq)
        ws.cell(row=8, column=col_seq).alignment = center_align

        # --- Row 9 Sub-Headers ---
        import string
        letters = string.ascii_uppercase # A, B, C...
        
        # Alpha Revisions
        if max_alpha > 0:
            for i in range(max_alpha):
                col = start_approval + i
                val = f"Rev {letters[i]}" if i < 26 else f"Rev {letters[25]}+{i}" 
                c = ws.cell(row=9, column=col, value=val)
                c.border = thin_border
                c.font = header_font
                c.alignment = center_align
                ws.column_dimensions[get_column_letter(col)].width = 12

        # Num Revisions
        if max_num > 0:
            for i in range(max_num):
                col = start_fab + i
                val = f"Rev {i}"
                c = ws.cell(row=9, column=col, value=val)
                c.border = thin_border
                c.font = header_font
                c.alignment = center_align
                ws.column_dimensions[get_column_letter(col)].width = 12

        # Widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions[get_column_letter(col_remarks)].width = 30
        ws.column_dimensions[get_column_letter(col_seq)].width = 15

        # Helper to categorize based on drawing data
        e_sheets = []
        shop_drawings = []
        part_drawings = []
        
        for d in drawing_log_data:
            d_type = str(d.get('drawing_type', '')).upper()
            d_no = str(d.get('drawing_no', '')).strip()
            d_no_upper = d_no.upper()
            description = str(d.get('description', '')).upper()
            
            # 1. Erection Drawings
            if d_type == 'ERECTION' or d_no_upper.startswith('E') or 'ERECTION' in description:
                e_sheets.append(d)
            else:
                # Heuristic for Shop vs Part
                # Shop (Assembly): Usually starts with B, C, G, HSS, AL, L (if lintel), etc.
                # Part: Usually starts with p, pl, a, bar, f, m (often lowercase in some naming, but here normalized)
                # We check Title keywords and Prefix patterns.
                
                is_part = False
                
                # Check for explicit Part patterns (starts with number then p, pl, a, w, m, f then number)
                # Regex: ^[0-9]*[a-zA-Z]+[0-9]+$ matches most Tekla marks (100B1, 100p1)
                
                # Strong indicators for Part (Use 'p' for small parts, 'a' for angles)
                # removed 'pl' from here, so 1PL1 goes to Shop. 'p' matches 'p1' but not 'pl1'
                if re.match(r'^\d*(p|a|f|m|bar|rod)\d+$', d_no, re.IGNORECASE):
                    is_part = True
                elif "ANGLE" in description and "LINTEL" not in description:
                    is_part = True
                
                # Strong indicators for Shop (Assemblies)
                # Beams, Columns, HSS, Stairs, Rails, Lintels, AND PLATES (PL)
                if "BEAM" in description or "COLUMN" in description or "GIRDER" in description or "HSS" in description or "PLATE" in description:
                    is_part = False
                elif re.match(r'^\d*(B|C|G|HSS|VB|HB|TR|STR|HR|AL|PL)\d+', d_no_upper):
                    is_part = False
                elif "LINTEL" in description: 
                    # Lintel can be tricky, usually assembly if "Lintel Angle"
                    is_part = False

                if is_part:
                    part_drawings.append(d)
                else:
                    shop_drawings.append(d)

        # Sort
        e_sheets.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))
        shop_drawings.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))
        part_drawings.sort(key=lambda x: self._natural_keys(x.get('drawing_no', '')))

        current_row = 10
        
        def write_log_section(title, items, start_sl):
            nonlocal current_row
            
            # Section Header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_seq)
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = header_font
            cell.fill = yellow_fill
            cell.alignment = center_align
            # Borders for header row
            for col in range(1, col_seq + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            
            current_row += 1
            
            for i, item in enumerate(items, start_sl):
                # Sl
                ws.cell(row=current_row, column=1, value=i).border = thin_border
                ws.cell(row=current_row, column=1).alignment = center_align
                
                # Sheet
                ws.cell(row=current_row, column=2, value=item.get('drawing_no', '')).border = thin_border
                ws.cell(row=current_row, column=2).alignment = center_align
                
                # Title
                ws.cell(row=current_row, column=3, value=item.get('description', '')).border = thin_border
                ws.cell(row=current_row, column=3).alignment = left_align
                
                # Revisions
                revs = item.get('revisions', [])
                item_alphas = {str(r['rev']).upper(): r['date'] for r in revs if str(r['rev']).upper() in letters}
                item_nums = {str(r['rev']): r['date'] for r in revs if str(r['rev']).isdigit()}
                
                # Fill Alpha Cols
                if max_alpha > 0:
                    for idx in range(max_alpha):
                        rev_char = letters[idx] # A, B
                        val = item_alphas.get(rev_char, '')
                        if val and '-' in str(val):
                             try: val = datetime.strptime(str(val).split()[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                             except: pass
                        c = ws.cell(row=current_row, column=start_approval+idx, value=val)
                        c.border = thin_border
                        c.alignment = center_align
                    
                # Fill Num Cols
                if max_num > 0:
                    for idx in range(max_num):
                        rev_num = str(idx) # "0", "1"
                        val = item_nums.get(rev_num, '')
                        if val and '-' in str(val):
                             try: val = datetime.strptime(str(val).split()[0], "%Y-%m-%d").strftime("%d/%m/%Y")
                             except: pass
                        c = ws.cell(row=current_row, column=start_fab+idx, value=val)
                        c.border = thin_border
                        c.alignment = center_align
                
                # Remarks - Get latest status or remarks from item
                # Transmittal logic uses remarks. Drawing log logic might use 'status' or 'remarks'
                rem = item.get('remarks', '')
                if not rem: rem = item.get('status', '')
                
                ws.cell(row=current_row, column=col_remarks, value=rem).border = thin_border
                ws.cell(row=current_row, column=col_remarks).alignment = left_align
                
                # Seq
                ws.cell(row=current_row, column=col_seq, value="").border = thin_border
                
                current_row += 1

        # Write Sections
        if e_sheets:
            write_log_section("E- PLAns", e_sheets, 1)
        
        if shop_drawings:
            write_log_section("SHOP DRAWINGS", shop_drawings, 1)
            
        if part_drawings:
            write_log_section("PART DRAWINGS", part_drawings, 1)

        wb.save(output_path)
        return output_path
